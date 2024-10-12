import random
import numpy as np
import pandas as pd
import datetime
from shapely.geometry import Polygon, Point
import plotly.graph_objects as go
import plotly.express as px
import os, sys
import plotly.offline
from dataclasses import dataclass, field
from typing import Dict, List
from alive_progress import alive_bar
from deap import base
from deap import creator
from deap import tools
from IPython.display import clear_output
import scipy.interpolate as interpolate
import time
import openpyxl
import discord
import nest_asyncio
import asyncio


def time_convert(date):
    fmt = "%d %b %Y %H:%M:%S.%f"
    try:
        t = datetime.datetime.strptime(date[:-3], fmt)
    except:
        t = datetime.datetime.strptime(date, fmt)
    return pd.Timestamp(year=t.year, month=t.month, day=t.day, hour = t.hour, minute = t.minute ,second=t.second, microsecond=t.microsecond)

def Create_Poly(filename):
    df = pd.read_csv(filename)
    l = []
    for i in range(len(df)):
        l.append((df['Lat'][i],df['Lon'][i]))
    return Polygon(l)

def get_ind(n_sats):
    df = pd.read_csv(f"../../Input_Files/Satellites_File_{n_sats}.txt")
    return [df["Per"].values[0],df["Inc"].values[0],max(df["Asc"]),len(np.unique(df["Loc"]))]

def plot_targets_and_polygon(poly,filename):
    df = pd.read_csv(filename)
    fig = go.Figure(go.Scattermapbox(
        mode = "markers",
        lon = df['Lon'],
        lat = df['Lat'],
        marker = {'size': 10}))

    fig.add_trace(go.Scattermapbox(
        mode = "lines",
        lon = np.array(poly.exterior.coords.xy)[0],
        lat = np.array(poly.exterior.coords.xy)[1],
        marker = {'size': 10}))

    fig.update_layout(
        margin ={'l':0,'t':0,'b':0,'r':0},
        mapbox = {
            'center': {'lon': 0, 'lat': 0},
            'style': "open-street-map",
            'center': {'lon': 0, 'lat': 0},
            'zoom': 0})
    return fig

def Pointing_File_Generator(filename,period):
    f = open(filename,"w")
    f.write("stk.v.12.1.1\nBegin\tAttitude\nNumberofAttitudePoints\t162\nSequence\t323\nRepeatPattern\n")
    for i in range(162):
        f.write(f'{period/162*(i+1)} {(i%9+1)*10-5} {(i//9+1)*10-5}\n')
    f.write('End Attitude')
    f.close()
class Optimizer:
    def __init__(self,stk_object,n_pop,n_gen,n_sats,weights = (7.0,-6.0,-1.0,-1.0)):
        self.stk_object = stk_object
        self.n_pop = n_pop
        self.n_gen = n_gen
        self.n_sats = n_sats
        creator.create("FitnessMax", base.Fitness, weights=weights)
        creator.create("Satellite", list, fitness=creator.FitnessMax)
        self.lower = [575,45,0,0,1]
        self.upper = [630,135,180,50,self.n_sats]

        # Registering variables to the satellite
        self.toolbox = base.Toolbox()
        self.toolbox.register("attr_alt", random.randint, self.lower[0], self.upper[0])
        self.toolbox.register("attr_inc", random.randint, self.lower[1], self.upper[1])
        self.toolbox.register("attr_initial_raan", random.randint, self.lower[2], self.upper[2])
        self.toolbox.register("attr_delta_raan", random.randint, self.lower[3], self.upper[3])
        self.toolbox.register("attr_num_planes", random.randint, self.lower[4], self.upper[4])

        # Registering satellite to the model
        self.toolbox.register("satellite", tools.initCycle, creator.Satellite,
                        (self.toolbox.attr_alt,
                        self.toolbox.attr_inc,
                        self.toolbox.attr_initial_raan,
                        self.toolbox.attr_delta_raan,
                        self.toolbox.attr_num_planes), n=1)

        # Registering tools for the algorithm
        self.toolbox.register("population", tools.initRepeat, list, self.toolbox.satellite)
        self.toolbox.register("evaluate", self.cost_function)
        self.toolbox.register("mate", tools.cxUniform,indpb=0.5)
        self.toolbox.register("mutate", tools.mutUniformInt, low=self.lower,up=self.upper, indpb=0.5)
        self.toolbox.register("select", tools.selTournament, tournsize=int(self.n_pop//2))
        self.stats = tools.Statistics(key=lambda ind: ind.fitness.values)
        self.stats.register("avg", np.mean, axis=0)
        self.stats.register("std", np.std, axis=0)
        self.stats.register("min", np.min, axis=0)
        self.stats.register("max", np.max, axis=0)

    def run(self,read=False,enable_print=False,file=False):
        self.enable_print = enable_print
        self.fits = []
        CXPB = 0.7;MUTPB=0.3
        g = 0
        # Creating a population to evolve
        pop = self.toolbox.population(n=self.n_pop)
        i = np.random.randint(0,self.n_pop)
        if read:
            for idx in range(4):
                pop[i][idx] = get_ind(self.n_sats)[idx]

        fitnesses = list(map(self.toolbox.evaluate, pop))
        for ind, fit in zip(pop, fitnesses):
            ind.fitness.values = fit
        hof = tools.HallOfFame(5)
        hof.update(pop)

        self.fits.append([ind.fitness.values for ind in pop])
        record = self.stats.compile(pop)

        clear_output(wait=False)
        print("-- Generation %i --" % g)
        print(pd.DataFrame(record))
        
        if file:
            file.write("Gen,Pop,Alt,Inc,Initial_Raan,Delta_Raan,Num_Planes,Avg_Percentage,Std_Percentage,Avg_Time,Std_Time,\n")
            for i in range(self.n_pop):
                file.write(f"{g},{i},")
                for idx in range(4):
                    file.write(f"{pop[i][idx]},")
                for fit in pop[i].fitness.getValues():
                    file.write(f"{fit},")
                file.write("\n")
        # Begin the evolution
        while g < self.n_gen:
            # clear_output(wait=True)
            g = g + 1

            # A new generation
            # Select the next generation individuals
            offspring = self.toolbox.select(pop, len(pop))
            # Clone the selected individuals
            offspring = list(map(self.toolbox.clone, offspring))
            # Apply crossover and mutation on the offspring
            for child1, child2 in zip(offspring[::2], offspring[1::2]):
                if random.random() < CXPB:
                    self.toolbox.mate(child1, child2)
                    del child1.fitness.values
                    del child2.fitness.values

            for mutant in offspring:
                if random.random() < MUTPB:
                    self.toolbox.mutate(mutant)
                    del mutant.fitness.values
            # Evaluate the individuals with an invalid fitness
            invalid_ind = [ind for ind in offspring if not ind.fitness.valid]
            fitnesses = map(self.toolbox.evaluate, invalid_ind)

            for ind, fit in zip(invalid_ind, fitnesses):
                ind.fitness.values = fit
            pop[:] = offspring
            hof.update(pop)

            self.fits.append([ind.fitness.values for ind in pop])
            record = self.stats.compile(pop)

            clear_output(wait=False)
            print("-- Generation %i --" % g)
            print(pd.DataFrame(record))
            
            if file:
                for i in range(self.n_pop):
                    file.write(f"{g},{i},")
                    for idx in range(5):
                        file.write(f"{pop[i][idx]},")
                    for fit in pop[i].fitness.getValues():
                        file.write(f"{fit},")
                    file.write("\n")
                
        return hof
    
    def cost_function(self,Individual=[0,0,0,0,0],write=True):
        if write:
            self.Load_Individual(Individual)
        self.stk_object.Satellite_Loader(f'../../Input_Files/Satellites_File_{self.n_sats}.txt')
        self.stk_object.Compute_AzEl(self.enable_print)
        percentages = np.array([100*np.count_nonzero(self.stk_object.target_bins[idx])/324 for idx in range(len(self.stk_object.targets))])
        times = np.array([self.stk_object.target_times[idx]/86400 for idx in range(len(self.stk_object.targets))])
        return np.average(percentages),np.std(percentages),np.average(times),np.std(times)
    
    def Load_Individual(self,Individual=[0,0,0,0,0]):
        Alt = Individual[0]
        Inc = Individual[1]
        initial_raan = Individual[2]
        delta_raan = Individual[3]
        num_planes = int(Individual[4])
        num_sats = self.n_sats
        Asc = initial_raan
        file = open(f"../../Input_Files/Satellites_File_{num_sats}.txt","w")
        file.write("Per,Apo,Inc,Asc,Loc\n")
        sats = num_sats*[1]
        planes = np.array_split(sats,num_planes)
        i=1
        for plane in planes:
            Loc = 0
            for sat in plane:
                file.write(f"{Alt},{Alt},{Inc},{round(Asc%360,4)},{round(Loc,4)}\n")
                if len(plane)>1: Loc += 360/len(plane)
            if len(planes)>1:Asc -= i*((-1)**(i))*delta_raan
            i+=1
        file.close()
        
def Interpolate(time,az,el):
    times = np.arange(time[0],time[-1],2.5)
    if max(el)>=60 and len(time)>3:
        az_t = interpolate.interpn(points=[time],values=np.array([np.unwrap(az,period=360)]).T,xi=times,method='pchip')[:,0]%360
        el_t = interpolate.interp1d(x=time,y=[el],kind='cubic')(times)[0]
    else:
        ans = interpolate.interp1d(x=time,y=[np.unwrap(az,period=360),el],kind='quadratic')(times).T
        az_t = ans[:,0]%360;el_t = ans[:,1]
    return times,az_t,el_t

def check_manueverability(previous_times,
                          previous_crossrange,
                          previous_alongrange,
                          new_time,
                          new_crossrange,
                          new_along_range,
                          slew_rate,
                          cone_angle):

    d_theta_2 = (new_crossrange**2+new_along_range**2)**0.5-cone_angle
    if d_theta_2 < 0:
        d_theta_2 = 0

    if len(previous_times)>0:
        d_theta_1 = (previous_crossrange**2+previous_alongrange**2)**0.5-cone_angle
        d_theta_1[d_theta_1<0] = 0

        d_time = np.abs(new_time-previous_times)

        return np.divide(d_theta_1+d_theta_2,d_time,out=1.1*slew_rate*np.ones_like(d_time),where=d_time!=0)<=slew_rate
    elif (slew_rate==0 and d_theta_2==0) or (slew_rate>0):
        return [[True]]
    else:
        return [[False]]

def get_best_available_access(satellite_specific_plan,bin_access_points,slew_rate,cone_angle,time_threshold):
    if len(bin_access_points)>0:
        for idx in range(len(bin_access_points)):
            previous_sat_accesses = satellite_specific_plan[int(bin_access_points[idx,3])]
            window = [i for i, t in enumerate(previous_sat_accesses["Time"]) if abs(t - bin_access_points[idx,0]) <= time_threshold]
            feasible = check_manueverability(np.array(previous_sat_accesses["Time"])[window],
                                             np.array(previous_sat_accesses["Cross Range"])[window],
                                             np.array(previous_sat_accesses["Along Range"])[window],
                                             bin_access_points[idx,0],
                                             bin_access_points[idx,1],
                                             bin_access_points[idx,2],
                                             slew_rate,
                                             cone_angle)
            if np.all(feasible):
                return bin_access_points[idx]
    return False

def dfs_to_excel(excel_file, sheet_name, n_sats, df2, df3):
    df1 = pd.read_csv(f'../../Input_Files/Satellites_File_{n_sats}.txt')
    workbook = openpyxl.load_workbook(excel_file)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

    for i, row in df1.iterrows():
        for j, key in enumerate(df1.columns):
            sheet.cell(row=1, column=j+1).value = key
            sheet.cell(row=i+2, column=j+1).value = row[key]
    for i, row in df2.iterrows():
        for j, key in enumerate(df2.columns):
            sheet.cell(row=1, column=j+len(df1.columns)+2).value = key
            sheet.cell(row=i+2, column=j+len(df1.columns)+2).value = row[key]

    for i, row in df3.iterrows():
        for j, key in enumerate(df3.columns):
            sheet.cell(row=i+15, column=j+1).value = row[key]

    workbook.save(excel_file)
    return 0

def Generate_Performance_Curve(file=r"H:/Shared drives/AERO 401 Project  L3Harris Team 1/Subteam Designs/OCD/Superstars.xlsx"):
    df = pd.read_excel(file)
    df.columns = df.iloc[0]
    df = df.drop(df.index[0])

    fig = go.Figure()
    max_targets = 0
    for column in df.columns[1:]:
        col_df = df[df[column]>0]
        
        if len(col_df['Targets']) > 0:
            max_targets = max(col_df['Targets']) if max_targets < max(col_df['Targets']) else max_targets

        fig.add_trace(go.Scatter(
            x=col_df['Targets'],
            y=col_df[column],
            mode='lines+markers',
            name=str(column)
        ))

    fig.add_trace(go.Scatter(
        x=[10,max_targets+10],
        y=[30,30],
        mode = 'lines',
        line=dict(dash='dash',
                color='red',
                width=3),
        name='Constraint'
    ))

    fig.update_layout(
        title='Constellation Performance Comparison',
        xaxis_title='Number of Targets',
        yaxis_title='Average Days to 100%',
        legend_title='Number of Satellites',
        template='plotly',
        height=600,
        width=1000
    )

    fig.add_annotation(
        x=max_targets/1.5,  # Place it at the end of the constraint line
        y=29,  # Y-coordinate on the line
        text="30 Day Duration Constraint",  # Text annotation
        showarrow=False,  # No arrow, just the text
        font=dict(size=18, color='red'),  # Font size
    )
    fig.show()
    return 0

def send_message_to_discord(channel_id, message, bot_token):
    nest_asyncio.apply()
    intents = discord.Intents.default()
    intents.message_content = True
    bot = discord.Client(intents=intents)
    async def send_message_and_exit():
        channel = bot.get_channel(channel_id)
        if channel is not None:
            await channel.send(message)
        else:
            print("Channel not found.")
        await bot.close()
    @bot.event
    async def on_ready():
        await send_message_and_exit()
        await bot.close()
    bot.run(bot_token)
    return 0